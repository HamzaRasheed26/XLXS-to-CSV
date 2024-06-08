import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from dateutil.parser import parse as parse_date
import pandas as pd
import os
import re

class CSVGeneratorApp:
    def __init__(self, root):  # Use double underscores
        self.root = root
        self.root.title("BS OwnerData XLXS -> CSV")
        
        self.excel_file_path = tk.StringVar()
        self.output_dir_path = tk.StringVar()

        self.create_widgets()

    def create_widgets(self): 
        self.root.grid_columnconfigure(0, weight=9)  # Column 0 weight (80%)
        self.root.grid_columnconfigure(1, weight=1)  # Column 1 weight (20%)

        # Excel file selection
        tk.Label(self.root, text="XLXS file from owner:", anchor='w').grid(row=0, column=0, padx=(20,0), pady=(50,0), sticky='w')
        tk.Entry(self.root, textvariable=self.excel_file_path, width=50).grid(row=1, column=0, padx=(20,0), pady=(0,10), sticky='we')
        tk.Button(self.root, text="...", command=self.browse_excel_file).grid(row=1, column=1, padx=(0, 10), pady=(0,10), sticky='w')

        # Output directory selection
        tk.Label(self.root, text="CSV folder for output files:").grid(row=2, column=0, padx=(20,0), pady=(30,0), sticky='w')
        tk.Entry(self.root, textvariable=self.output_dir_path, width=50).grid(row=3, column=0, padx=(20,0), pady=(0,10), sticky='we')
        tk.Button(self.root, text="...", command=self.browse_output_directory).grid(row=3, column=1, padx=(0, 10), pady=(0,10), sticky='w')

        # Process button
        tk.Button(self.root, text="Process", command=self.process_files).grid(row=4, column=0, columnspan=2, pady=20)

        # Result display
        self.result_text = tk.Text(self.root, height=10, width=80)
        self.result_text.grid(row=5, column=0, columnspan=2, padx=10, pady=10, sticky='we')
    
        # Process button
        tk.Button(self.root, text="Close", command=self.close).grid(row=6, column=0, columnspan=2, pady=20, sticky='e', padx=10)

    def close(self):
        # Close the window
        self.root.destroy()

    def browse_excel_file(self):
        # Open file dialog to select Excel file
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if file_path:
            self.excel_file_path.set(file_path)

    def browse_output_directory(self):
        # Open file dialog to select output directory
        dir_path = filedialog.askdirectory()
        if dir_path:
            self.output_dir_path.set(dir_path)

    def process_files(self):
        # Clear the result text box
        self.result_text.delete(1.0, tk.END)
        
        # Get the paths
        excel_file = self.excel_file_path.get()
        output_dir = self.output_dir_path.get()
        
        # Check if paths are valid
        if not excel_file:
            self.result_text.insert(tk.END, "No Excel file selected\n")
            return

        # Check if output directory exists
        if not output_dir:
            self.result_text.insert(tk.END, "No folder selected\n")
            return

        # Check if excel file is readable
        try:
            wb = openpyxl.load_workbook(excel_file)
        except Exception as e:
            print(e)
            self.result_text.insert(tk.END, "Excel file not readable\n")
            return

        # Process the sheets in the workbook and generate CSV files
        for sheet in wb:
            try:
                if not self.validate_sheet(sheet): # Validate the sheet
                    return
                check = True
                data, check = self.extract_data(sheet) # Extract data from the sheet
                if not check:
                    return
                csv_filename = self.generate_csv(sheet, data, output_dir) # Generate CSV file
                self.result_text.insert(tk.END, f"CSV generated for sheet '{sheet.title}': {csv_filename}\n") # Display success message
            except Exception as e:
                # Display error message
                self.result_text.insert(tk.END, f"Error processing sheet '{sheet.title}': {e}\n")


    def generate_csv(self, sheet, data, output_dir):
        # Extract necessary information for filename
        sheet_name = sheet['C2'].value
        court_id = sheet['C3'].value
        from_date = sheet['D5'].value
        till_date = sheet['F5'].value

        # Create the filename
        filename = f"{sheet_name}-court {court_id}-{from_date}--{till_date}.csv"
        output_path = os.path.join(output_dir, filename)

        # Create DataFrame and save to CSV
        df = pd.DataFrame(data)
        df.to_csv(output_path, index=False, encoding='utf-8')

        return output_path

    def validate_sheet(self, sheet):
        # Check cell C1 for a valid integer > 0
        court_id = sheet['C3'].value
        if not isinstance(court_id, int) or court_id <= 0:
            tab_name = sheet.title
            self.result_text.insert(tk.END, f"Invalid Court ID on tab {tab_name}\n")
            return False

        # Check cell D5 for a valid date
        from_date = sheet['D5'].value
        try:
            parsed_date = parse_date(from_date, dayfirst=True)
        except (ValueError, TypeError):
            tab_name = sheet.title
            self.result_text.insert(tk.END, f"Invalid From date on tab {tab_name}\n")
            return False

        # Check cell F5 for a valid date
        till_date = sheet['F5'].value
        try:
            parsed_date = parse_date(till_date, dayfirst=True)
        except (ValueError, TypeError):
            tab_name = sheet.title
            self.result_text.insert(tk.END, f"Invalid Till date on tab {tab_name}\n")
            return False
        return True

    def extract_latest_booking_time(self, text):
        # Extract latest booking time
        match = re.search(r'\b(\d{1,2}:\d{2}\s*[apAP][mM])\b', text)
        if match:
            return match.group(1)
        return None

    def extract_data(self, sheet):
        # Extracting from and till date
        from_date_str = sheet['D5'].value
        till_date_str = sheet['F5'].value
        from_date = parse_date(from_date_str, dayfirst=True)
        till_date = parse_date(till_date_str, dayfirst=True)

        # Extracting latest booking time
        # booking time form Monday to Friday
        raw_booking_time = sheet['C6'].value
        booking_time_M_F = self.extract_latest_booking_time(raw_booking_time).lower().replace(" am", "").replace(" pm", "")
        # booking time on Saturday to Sunday
        raw_booking_time = sheet['C7'].value
        booking_time_Ss = self.extract_latest_booking_time(raw_booking_time).lower().replace(" am", "").replace(" pm", "")
        if booking_time_M_F is None or booking_time_Ss is None:
            tab_name = sheet.title
            self.result_text.insert(tk.END, f"Invalid booking time on tab {tab_name}\n")
            return [], False

        # Extract all possible dates from to till date
        period_global = []
        period_global = self.possible_dates(from_date, till_date)
        
        # Extracting nbr of slots
        slots = self.nbr_of_slots(sheet)

        # Extracting week day hours
        week_times = []
        week_times, check = self.extract_week_times(sheet, slots)
        if not check:
            return [], False

        # Extracting exceptions dates for the selected period:
        exception_period = []
        check = True
        exception_period, check = self.extract_exception_dates(sheet)
        if not check:
            return [], False

        # Excluding the exception dates from the period
        period_global = self.exclude_exception_dates(period_global, exception_period)

        # Making the combination of all the data
        commercial_hours = []
        commercial_hours = self.making_combination(period_global, week_times, booking_time_M_F, booking_time_Ss)

        return commercial_hours, True

    def possible_dates(self, from_date, till_date):
        # Extract all possible dates from to till date
        period_global = []
        current_date = from_date
        while current_date <= till_date:
            period_global.append({
                "day_of_week": current_date.strftime('%A'),
                "date": current_date.strftime('%d/%m/%Y')
            })
            current_date += timedelta(days=1)
        return period_global

    def nbr_of_slots(self, sheet):
        # Extracting nbr of slots
        slots = 0
        col = 3
        while True:
            header_start = sheet.cell(row=10, column=col).value
            header_end = sheet.cell(row=10, column=col+1).value
            if header_start.lower() == "start time" and header_end.lower() == "end time":
                slots += 1
                col += 2
            else:
                break
        return slots

    def extract_week_times(self, sheet, slots):
        # Extracting week day hours
        week_times = []
        row = 11
        try:
            for i in range(7):
                col = 3
                for j in range(slots):
                    start_time = sheet.cell(row=row, column=col).value
                    end_time = sheet.cell(row=row, column=col+1).value
                    if start_time and end_time:
                        week_times.append({
                            "day": sheet.cell(row=row, column=2).value,
                            "start_time": start_time,
                            "end_time": end_time
                        })
                    col += 2
                row += 1
        except:
            tab_name = sheet.title
            self.result_text.insert(tk.END, f"Invalid week day hours on tab {tab_name}\n")
            return [], False
        return week_times, True

    def extract_exception_dates(self, sheet):
        # Extracting exceptions dates for the selected period:
        exception_period = []

        row = 21
        while True:
            cell_value = sheet.cell(row=row, column=9).value  # Column I is the 9th column
            if cell_value:
                if '-' in cell_value or 'od' in cell_value:  # It's a range
                    parts = re.split(r'[-\s*od\s*]', cell_value)
                    if len(parts) == 2:
                        try:
                            start_date = parse_date(parts[0].strip(), dayfirst=True)
                            end_date = parse_date(parts[1].strip(), dayfirst=True)
                            current_date = start_date
                            while current_date <= end_date:
                                exception_period.append(current_date.strftime('%d/%m/%Y'))
                                current_date += timedelta(days=1)
                        except ValueError:
                            tab_name = sheet.title
                            self.result_text.insert(tk.END, f"Invalid exception date on tab {tab_name}, cell {sheet.cell(row=row, column=9).coordinate}\n")
                            return [], False
                            # raise ValueError(f"Invalid exception date on tab {tab_name}, cell {sheet.cell(row=row, column=9).coordinate}")
                else:  # It's a single date
                    try:
                        exception_date = parse_date(cell_value.strip(), dayfirst=True)
                        exception_period.append(exception_date.strftime('%d/%m/%Y'))
                    except ValueError:
                        tab_name = sheet.title
                        self.result_text.insert(tk.END, f"Invalid exception date on tab {tab_name}, cell {sheet.cell(row=row, column=9).coordinate}\n")
                        return [], False
                        # raise ValueError(f"Invalid exception date on tab {tab_name}, cell {sheet.cell(row=row, column=9).coordinate}")
            else:
                break
            row += 1

        return exception_period, True

    def exclude_exception_dates(self, period_global, exception_period):
        # Excluding the exception dates from the period
        exception_set = set(exception_period)
        period_global[:] = [day_info for day_info in period_global if day_info["date"] not in exception_set]
        return period_global

    def making_combination(self, period_global, week_times, booking_time_M_F, booking_time_Ss):
        # Making the combination of all the data
        commercial_hours = []
        for day in period_global:
            for time in week_times:
                if day['day_of_week'].lower() == time['day'].lower():
                    if day['day_of_week'].lower() in ['saturday', 'sunday']:
                        book_time = booking_time_Ss
                    else:
                        book_time = booking_time_M_F

                    commercial_hours.append({
                        "Date": day['date'],
                        "Start Time": time['start_time'],
                        "End Time": time['end_time'],
                        "Latest Booking Date": day['date'],
                        "Latest Booking Time": book_time,
                        "Rate, PLN": "-",
                        "Description": "Management",
                        "User Id": "",
                        "Commission enabled": "",
                        "Commission in percent": "",
                        "Commission": ""
                    })
        return commercial_hours

if __name__ == "__main__":
    root = tk.Tk()
    app = CSVGeneratorApp(root)
    root.mainloop()